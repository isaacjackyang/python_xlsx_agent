from __future__ import annotations

import re
from typing import Any

from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import range_boundaries

from xl_agent_core.core.contracts import FormulaTraceLink, FormulaTraceResult, WarningMessage
from xl_agent_core.core.loader import get_sheet_pair, load_workbook_bundle, normalize_ref


class FormulaService:
    def formula_map(self, path: str, sheet: str) -> dict[str, list[dict[str, str]]]:
        bundle = load_workbook_bundle(path)
        worksheet, _ = get_sheet_pair(bundle, sheet)
        mapping: dict[str, list[dict[str, str]]] = {}

        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    mapping[cell.coordinate] = [
                        {"sheet": ref_sheet, "ref": ref_value, "kind": ref_kind}
                        for ref_sheet, ref_value, ref_kind in self._extract_references(cell.value, worksheet.title)
                    ]

        return mapping

    def formula_trace(self, path: str, sheet: str, cell: str, direction: str) -> FormulaTraceResult:
        bundle = load_workbook_bundle(path)
        worksheet, _ = get_sheet_pair(bundle, sheet)
        normalized_cell = normalize_ref(cell)
        target = worksheet[normalized_cell]
        formula = target.value if isinstance(target.value, str) and target.value.startswith("=") else None
        warnings: list[WarningMessage] = []

        if direction == "precedents":
            if formula is None:
                warnings.append(
                    WarningMessage(
                        code="cell_is_not_formula",
                        message=f"Cell {normalized_cell} does not contain a formula.",
                    ),
                )
                links: list[FormulaTraceLink] = []
            else:
                links = [
                    FormulaTraceLink(
                        sheet=ref_sheet,
                        ref=ref_value,
                        kind=ref_kind,
                        formula_cell=normalized_cell,
                        formula=formula,
                    )
                    for ref_sheet, ref_value, ref_kind in self._extract_references(formula, worksheet.title)
                ]
        elif direction == "dependents":
            links = self._find_dependents(bundle.formulas, sheet, normalized_cell)
            if not links:
                warnings.append(
                    WarningMessage(
                        code="no_dependents_found",
                        message=f"No dependent formulas were found for {sheet}!{normalized_cell}.",
                    ),
                )
        else:
            raise ValueError("direction must be 'precedents' or 'dependents'.")

        return FormulaTraceResult(
            workbook_path=str(bundle.path),
            sheet=sheet,
            cell=normalized_cell,
            direction=direction,
            formula=formula,
            links=links,
            warnings=warnings,
        )

    def _find_dependents(self, workbook: Any, target_sheet: str, target_cell: str) -> list[FormulaTraceLink]:
        links: list[FormulaTraceLink] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                        continue
                    references = self._extract_references(cell.value, worksheet.title)
                    for ref_sheet, ref_value, ref_kind in references:
                        if self._reference_matches_target(ref_sheet, ref_value, ref_kind, target_sheet, target_cell, worksheet.title):
                            links.append(
                                FormulaTraceLink(
                                    sheet=worksheet.title,
                                    ref=cell.coordinate,
                                    kind="dependent_formula",
                                    formula_cell=cell.coordinate,
                                    formula=cell.value,
                                ),
                            )
                            break
        return links

    def _reference_matches_target(
        self,
        reference_sheet: str,
        reference_value: str,
        reference_kind: str,
        target_sheet: str,
        target_cell: str,
        current_formula_sheet: str,
    ) -> bool:
        effective_sheet = reference_sheet or current_formula_sheet
        if effective_sheet != target_sheet:
            return False

        if reference_kind == "cell":
            return normalize_ref(reference_value) == normalize_ref(target_cell)
        if reference_kind == "range":
            return self._range_contains(reference_value, target_cell)
        return False

    def _range_contains(self, range_ref: str, target_cell: str) -> bool:
        min_col, min_row, max_col, max_row = range_boundaries(normalize_ref(range_ref))
        target_match = re.fullmatch(r"([A-Z]+)([0-9]+)", normalize_ref(target_cell))
        if not target_match:
            return False

        from openpyxl.utils.cell import column_index_from_string

        column_name, row_number = target_match.groups()
        column_index = column_index_from_string(column_name)
        row_index = int(row_number)
        return min_col <= column_index <= max_col and min_row <= row_index <= max_row

    def _extract_references(self, formula: str, current_sheet: str) -> list[tuple[str, str, str]]:
        tokenizer = Tokenizer(formula)
        refs: list[tuple[str, str, str]] = []

        for token in tokenizer.items:
            if token.type != "OPERAND" or token.subtype != "RANGE":
                continue

            raw = token.value.replace("$", "")
            ref_sheet = current_sheet
            ref_value = raw
            if "!" in raw:
                sheet_part, ref_value = raw.rsplit("!", 1)
                ref_sheet = sheet_part.strip("'")

            if self._looks_like_cell(ref_value):
                ref_kind = "cell"
            elif self._looks_like_range(ref_value):
                ref_kind = "range"
            elif "[" in ref_value and "]" in ref_value:
                ref_kind = "structured_ref"
            else:
                ref_kind = "name"

            if ref_kind in {"cell", "range"}:
                ref_value = normalize_ref(ref_value)

            refs.append((ref_sheet, ref_value, ref_kind))

        return refs

    def _looks_like_cell(self, value: str) -> bool:
        return bool(re.fullmatch(r"[A-Z]+[0-9]+", normalize_ref(value)))

    def _looks_like_range(self, value: str) -> bool:
        return bool(re.fullmatch(r"[A-Z]+[0-9]+:[A-Z]+[0-9]+", normalize_ref(value)))
