from __future__ import annotations

from dataclasses import dataclass
from typing import Any

from openpyxl.utils.cell import get_column_letter

from xl_agent_core.core.contracts import RegionCandidate
from xl_agent_core.core.loader import is_populated, load_workbook_bundle


@dataclass(slots=True)
class _RegionBox:
    start_row: int
    end_row: int
    start_col: int
    end_col: int

    @property
    def width(self) -> int:
        return self.end_col - self.start_col + 1

    @property
    def height(self) -> int:
        return self.end_row - self.start_row + 1

    @property
    def range_ref(self) -> str:
        return (
            f"{get_column_letter(self.start_col)}{self.start_row}:"
            f"{get_column_letter(self.end_col)}{self.end_row}"
        )


class RegionDetector:
    def detect(self, path: str, sheet: str) -> list[RegionCandidate]:
        bundle = load_workbook_bundle(path)
        worksheet = bundle.formulas[sheet]

        occupied = self._occupied_coordinates(worksheet)
        if not occupied:
            return []

        row_groups = self._group_integers(sorted({row for row, _ in occupied}))
        candidates: list[RegionCandidate] = []
        counter = 1

        for row_group in row_groups:
            cols_in_group = sorted({col for row, col in occupied if row in row_group})
            for col_group in self._group_integers(cols_in_group):
                box = _RegionBox(
                    start_row=min(row_group),
                    end_row=max(row_group),
                    start_col=min(col_group),
                    end_col=max(col_group),
                )
                trimmed = self._trim_box(worksheet, box)
                if trimmed is None:
                    continue

                non_empty = self._count_non_empty(worksheet, trimmed)
                area = trimmed.width * trimmed.height
                density = non_empty / area if area else 0.0
                if non_empty == 0:
                    continue

                candidates.append(
                    RegionCandidate(
                        sheet=sheet,
                        label=f"region_{counter}",
                        range_ref=trimmed.range_ref,
                        header_row=self._guess_header_row(worksheet, trimmed),
                        non_empty_cells=non_empty,
                        density=round(density, 3),
                        width=trimmed.width,
                        height=trimmed.height,
                        notes=self._notes_for_box(worksheet, trimmed),
                    ),
                )
                counter += 1

        candidates.sort(key=lambda item: (item.header_row or 10**9, item.range_ref))
        return candidates

    @staticmethod
    def select_best(candidates: list[RegionCandidate]) -> RegionCandidate:
        if not candidates:
            raise ValueError("No candidate regions were detected.")
        return max(
            candidates,
            key=lambda item: (
                item.non_empty_cells,
                item.density,
                item.height > 1,
                item.width > 1,
            ),
        )

    def _occupied_coordinates(self, worksheet: Any) -> list[tuple[int, int]]:
        occupied: list[tuple[int, int]] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if is_populated(cell.value):
                    occupied.append((cell.row, cell.column))
        return occupied

    def _group_integers(self, numbers: list[int]) -> list[list[int]]:
        if not numbers:
            return []

        groups: list[list[int]] = [[numbers[0]]]
        for number in numbers[1:]:
            if number == groups[-1][-1] + 1:
                groups[-1].append(number)
            else:
                groups.append([number])
        return groups

    def _trim_box(self, worksheet: Any, box: _RegionBox) -> _RegionBox | None:
        start_row = box.start_row
        end_row = box.end_row
        start_col = box.start_col
        end_col = box.end_col

        while start_row <= end_row and self._row_is_empty(worksheet, start_row, start_col, end_col):
            start_row += 1
        while end_row >= start_row and self._row_is_empty(worksheet, end_row, start_col, end_col):
            end_row -= 1
        while start_col <= end_col and self._column_is_empty(worksheet, start_col, start_row, end_row):
            start_col += 1
        while end_col >= start_col and self._column_is_empty(worksheet, end_col, start_row, end_row):
            end_col -= 1

        if start_row > end_row or start_col > end_col:
            return None

        return _RegionBox(start_row=start_row, end_row=end_row, start_col=start_col, end_col=end_col)

    def _row_is_empty(self, worksheet: Any, row: int, start_col: int, end_col: int) -> bool:
        for column in range(start_col, end_col + 1):
            if is_populated(worksheet.cell(row=row, column=column).value):
                return False
        return True

    def _column_is_empty(self, worksheet: Any, column: int, start_row: int, end_row: int) -> bool:
        for row in range(start_row, end_row + 1):
            if is_populated(worksheet.cell(row=row, column=column).value):
                return False
        return True

    def _count_non_empty(self, worksheet: Any, box: _RegionBox) -> int:
        count = 0
        for row in range(box.start_row, box.end_row + 1):
            for column in range(box.start_col, box.end_col + 1):
                if is_populated(worksheet.cell(row=row, column=column).value):
                    count += 1
        return count

    def _guess_header_row(self, worksheet: Any, box: _RegionBox) -> int | None:
        first_row_values = [
            worksheet.cell(row=box.start_row, column=column).value
            for column in range(box.start_col, box.end_col + 1)
        ]
        text_like = [value for value in first_row_values if isinstance(value, str) and value.strip()]
        return box.start_row if text_like else None

    def _notes_for_box(self, worksheet: Any, box: _RegionBox) -> list[str]:
        notes: list[str] = []

        hidden_rows = [row for row in range(box.start_row, box.end_row + 1) if worksheet.row_dimensions[row].hidden]
        if hidden_rows:
            notes.append("contains_hidden_rows")

        hidden_columns = [
            get_column_letter(column)
            for column in range(box.start_col, box.end_col + 1)
            if worksheet.column_dimensions[get_column_letter(column)].hidden
        ]
        if hidden_columns:
            notes.append("contains_hidden_columns")

        merged_ranges = [merged.coord for merged in worksheet.merged_cells.ranges if self._intersects(box, merged)]
        if merged_ranges:
            notes.append("contains_merged_cells")

        if worksheet.auto_filter.ref and self._range_strings_overlap(box.range_ref, worksheet.auto_filter.ref):
            notes.append("intersects_filter")

        return notes

    def _intersects(self, box: _RegionBox, merged_range: Any) -> bool:
        return not (
            merged_range.max_row < box.start_row
            or merged_range.min_row > box.end_row
            or merged_range.max_col < box.start_col
            or merged_range.min_col > box.end_col
        )

    def _range_strings_overlap(self, left: str, right: str) -> bool:
        from openpyxl.utils.cell import range_boundaries

        left_min_col, left_min_row, left_max_col, left_max_row = range_boundaries(left)
        right_min_col, right_min_row, right_max_col, right_max_row = range_boundaries(right)
        return not (
            right_max_row < left_min_row
            or right_min_row > left_max_row
            or right_max_col < left_min_col
            or right_min_col > left_max_col
        )
