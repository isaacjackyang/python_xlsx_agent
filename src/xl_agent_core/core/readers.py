from __future__ import annotations

import re
from collections import Counter

from openpyxl.utils.cell import get_column_letter, range_boundaries

from xl_agent_core.core.contracts import (
    CellInspection,
    LayoutReadResult,
    TableReadResult,
    TableSchemaColumn,
    WarningMessage,
)
from xl_agent_core.core.layout import inspect_layout
from xl_agent_core.core.loader import display_value, get_sheet_pair, load_workbook_bundle, normalize_ref
from xl_agent_core.core.regions import RegionDetector


class ReadService:
    def __init__(self) -> None:
        self._regions = RegionDetector()

    def read_table(self, path: str, sheet: str, region: str | None = None) -> TableReadResult:
        bundle = load_workbook_bundle(path)
        formula_ws, value_ws = get_sheet_pair(bundle, sheet)
        resolved_region = self._resolve_region(path, sheet, region)
        min_col, min_row, max_col, max_row = range_boundaries(normalize_ref(resolved_region))

        header_row_values = []
        for column in range(min_col, max_col + 1):
            formula_cell = formula_ws.cell(row=min_row, column=column)
            value_cell = value_ws.cell(row=min_row, column=column)
            header_row_values.append(display_value(formula_cell, value_cell))

        header = self._normalize_header(header_row_values)
        schema = [
            TableSchemaColumn(
                name=column_name,
                ref=f"{get_column_letter(min_col + index)}{min_row}",
                index=index,
            )
            for index, column_name in enumerate(header)
        ]

        rows: list[dict[str, object]] = []
        grid: list[list[object]] = []
        for row in range(min_row + 1, max_row + 1):
            row_values = []
            for column in range(min_col, max_col + 1):
                formula_cell = formula_ws.cell(row=row, column=column)
                value_cell = value_ws.cell(row=row, column=column)
                row_values.append(display_value(formula_cell, value_cell))
            grid.append(row_values)
            rows.append({header[index]: row_values[index] for index in range(len(header))})

        _, layout_warnings = inspect_layout(formula_ws, value_ws, resolved_region)
        warnings = list(layout_warnings)
        if max_row <= min_row:
            warnings.append(
                WarningMessage(
                    code="single_row_region",
                    message="Region has no data rows below the header row.",
                ),
            )

        return TableReadResult(
            workbook_path=str(bundle.path),
            sheet=sheet,
            region=resolved_region,
            header=header,
            schema=schema,
            rows=rows,
            grid=grid,
            row_count=len(rows),
            column_count=len(header),
            warnings=warnings,
        )

    def read_layout(self, path: str, sheet: str, range_ref: str) -> LayoutReadResult:
        bundle = load_workbook_bundle(path)
        formula_ws, value_ws = get_sheet_pair(bundle, sheet)
        cells, warnings = inspect_layout(formula_ws, value_ws, range_ref)
        min_col, min_row, max_col, max_row = range_boundaries(normalize_ref(range_ref))

        merged_ranges = [
            merged.coord
            for merged in formula_ws.merged_cells.ranges
            if not (
                merged.max_row < min_row
                or merged.min_row > max_row
                or merged.max_col < min_col
                or merged.min_col > max_col
            )
        ]
        hidden_rows = [row for row in range(min_row, max_row + 1) if formula_ws.row_dimensions[row].hidden]
        hidden_columns = [
            get_column_letter(column)
            for column in range(min_col, max_col + 1)
            if formula_ws.column_dimensions[get_column_letter(column)].hidden
        ]

        return LayoutReadResult(
            workbook_path=str(bundle.path),
            sheet=sheet,
            range_ref=normalize_ref(range_ref),
            cells=cells,
            merged_ranges=merged_ranges,
            hidden_rows=hidden_rows,
            hidden_columns=hidden_columns,
            filter_ref=formula_ws.auto_filter.ref or None,
            warnings=warnings,
        )

    def read_cells(self, path: str, sheet: str, refs: list[str]) -> list[CellInspection]:
        bundle = load_workbook_bundle(path)
        formula_ws, value_ws = get_sheet_pair(bundle, sheet)
        inspections: list[CellInspection] = []

        for ref in self._expand_refs(refs):
            row, column = self._ref_to_tuple(ref)
            formula_cell = formula_ws.cell(row=row, column=column)
            value_cell = value_ws.cell(row=row, column=column)
            merged_range = self._merged_range_for_ref(formula_ws, ref)
            inspections.append(
                CellInspection(
                    sheet=sheet,
                    ref=ref,
                    value=value_cell.value,
                    displayed_value=display_value(formula_cell, value_cell),
                    formula=formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None,
                    data_type=formula_cell.data_type,
                    is_merged=merged_range is not None,
                    merged_range=merged_range,
                    row_hidden=bool(formula_ws.row_dimensions[row].hidden),
                    column_hidden=bool(formula_ws.column_dimensions[get_column_letter(column)].hidden),
                ),
            )

        return inspections

    def _resolve_region(self, path: str, sheet: str, region: str | None) -> str:
        if region is None or region.lower() == "auto":
            candidates = self._regions.detect(path, sheet)
            return self._regions.select_best(candidates).range_ref

        normalized = normalize_ref(region)
        if self._looks_like_range(normalized):
            return normalized

        candidates = self._regions.detect(path, sheet)
        for candidate in candidates:
            if candidate.label.lower() == region.lower():
                return candidate.range_ref

        raise ValueError(f"Unknown region '{region}'.")

    def _normalize_header(self, values: list[object]) -> list[str]:
        base_names = []
        for index, value in enumerate(values, start=1):
            if value is None:
                base_names.append(f"column_{index}")
            else:
                text = str(value).strip()
                base_names.append(text or f"column_{index}")

        counts = Counter(base_names)
        seen: Counter[str] = Counter()
        output: list[str] = []
        for name in base_names:
            seen[name] += 1
            if counts[name] == 1:
                output.append(name)
            else:
                output.append(f"{name}_{seen[name]}")
        return output

    def _looks_like_range(self, ref: str) -> bool:
        return bool(re.fullmatch(r"[A-Z]+[0-9]+(:[A-Z]+[0-9]+)?", ref))

    def _expand_refs(self, refs: list[str]) -> list[str]:
        expanded: list[str] = []
        for ref in refs:
            normalized = normalize_ref(ref)
            if ":" not in normalized:
                expanded.append(normalized)
                continue
            min_col, min_row, max_col, max_row = range_boundaries(normalized)
            for row in range(min_row, max_row + 1):
                for column in range(min_col, max_col + 1):
                    expanded.append(f"{get_column_letter(column)}{row}")
        return expanded

    def _ref_to_tuple(self, ref: str) -> tuple[int, int]:
        match = re.fullmatch(r"([A-Z]+)([0-9]+)", normalize_ref(ref))
        if not match:
            raise ValueError(f"Invalid cell reference '{ref}'.")

        from openpyxl.utils.cell import column_index_from_string

        column_name, row_number = match.groups()
        return int(row_number), column_index_from_string(column_name)

    def _merged_range_for_ref(self, worksheet, ref: str) -> str | None:
        normalized = normalize_ref(ref)
        for merged_range in worksheet.merged_cells.ranges:
            if normalized in merged_range:
                return merged_range.coord
        return None
