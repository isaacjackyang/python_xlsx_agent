from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string


def is_populated(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str) and value.strip() == "":
        return False
    return True


@dataclass(slots=True)
class WorkbookBundle:
    path: Path
    formulas: Any
    values: Any


def load_workbook_bundle(path: str | Path) -> WorkbookBundle:
    workbook_path = Path(path).expanduser().resolve()
    formulas = load_workbook(workbook_path, data_only=False)
    values = load_workbook(workbook_path, data_only=True)
    return WorkbookBundle(path=workbook_path, formulas=formulas, values=values)


def get_sheet_pair(bundle: WorkbookBundle, sheet: str):
    if sheet not in bundle.formulas.sheetnames:
        raise ValueError(f"Sheet '{sheet}' was not found.")
    return bundle.formulas[sheet], bundle.values[sheet]


def display_value(formula_cell: Any, value_cell: Any) -> Any:
    if value_cell.value is not None:
        return value_cell.value
    return formula_cell.value


def normalize_ref(ref: str) -> str:
    return ref.replace("$", "").upper()


def coordinate_to_tuple(ref: str) -> tuple[int, int]:
    column_name, row_index = coordinate_from_string(normalize_ref(ref))
    from openpyxl.utils.cell import column_index_from_string

    return row_index, column_index_from_string(column_name)
