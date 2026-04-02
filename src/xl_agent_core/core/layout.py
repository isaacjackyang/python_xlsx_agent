from __future__ import annotations

from typing import Any

from openpyxl.utils.cell import get_column_letter, range_boundaries

from xl_agent_core.core.contracts import LayoutCell, WarningMessage
from xl_agent_core.core.loader import display_value, normalize_ref


def inspect_layout(formula_ws: Any, value_ws: Any, range_ref: str) -> tuple[list[LayoutCell], list[WarningMessage]]:
    min_col, min_row, max_col, max_row = range_boundaries(normalize_ref(range_ref))
    merged_lookup = _merged_lookup(formula_ws)
    cells: list[LayoutCell] = []
    warnings: list[WarningMessage] = []

    hidden_rows = [row for row in range(min_row, max_row + 1) if formula_ws.row_dimensions[row].hidden]
    if hidden_rows:
        warnings.append(
            WarningMessage(
                code="hidden_rows_in_range",
                message=f"Range contains hidden rows: {hidden_rows}",
            ),
        )

    hidden_columns = [
        get_column_letter(column)
        for column in range(min_col, max_col + 1)
        if formula_ws.column_dimensions[get_column_letter(column)].hidden
    ]
    if hidden_columns:
        warnings.append(
            WarningMessage(
                code="hidden_columns_in_range",
                message=f"Range contains hidden columns: {hidden_columns}",
            ),
        )

    merged_ranges = [
        coord
        for coord, bounds in merged_lookup.items()
        if _bounds_overlap(bounds, (min_col, min_row, max_col, max_row))
    ]
    merged_ranges = sorted(set(merged_ranges))
    if merged_ranges:
        warnings.append(
            WarningMessage(
                code="merged_cells_in_range",
                message=f"Range intersects merged cells: {merged_ranges}",
            ),
        )

    if formula_ws.auto_filter.ref and _range_overlap(formula_ws.auto_filter.ref, range_ref):
        warnings.append(
            WarningMessage(
                code="filter_overlap",
                message=f"Range overlaps auto-filter region {formula_ws.auto_filter.ref}.",
            ),
        )

    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            formula_cell = formula_ws.cell(row=row, column=column)
            value_cell = value_ws.cell(row=row, column=column)
            ref = f"{get_column_letter(column)}{row}"
            merged_range = _merged_range_for_ref(formula_ws, ref)
            cells.append(
                LayoutCell(
                    sheet=formula_ws.title,
                    ref=ref,
                    row=row,
                    column=column,
                    value=value_cell.value,
                    displayed_value=display_value(formula_cell, value_cell),
                    formula=formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None,
                    data_type=formula_cell.data_type,
                    is_merged=merged_range is not None,
                    merged_range=merged_range,
                    row_hidden=bool(formula_ws.row_dimensions[row].hidden),
                    column_hidden=bool(formula_ws.column_dimensions[get_column_letter(column)].hidden),
                    style={
                        "number_format": formula_cell.number_format,
                        "font_bold": bool(formula_cell.font and formula_cell.font.bold),
                        "font_italic": bool(formula_cell.font and formula_cell.font.italic),
                        "horizontal_alignment": getattr(formula_cell.alignment, "horizontal", None),
                        "fill_color": getattr(getattr(formula_cell.fill, "fgColor", None), "rgb", None),
                    },
                ),
            )

    return cells, warnings


def _merged_lookup(worksheet: Any) -> dict[str, tuple[int, int, int, int]]:
    lookup: dict[str, tuple[int, int, int, int]] = {}
    for merged_range in worksheet.merged_cells.ranges:
        lookup[merged_range.coord] = (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )
    return lookup


def _merged_range_for_ref(worksheet: Any, ref: str) -> str | None:
    normalized = normalize_ref(ref)
    for merged_range in worksheet.merged_cells.ranges:
        if normalized in merged_range:
            return merged_range.coord
    return None


def _bounds_overlap(left: tuple[int, int, int, int], right: tuple[int, int, int, int]) -> bool:
    left_min_col, left_min_row, left_max_col, left_max_row = left
    right_min_col, right_min_row, right_max_col, right_max_row = right
    return not (
        right_max_row < left_min_row
        or right_min_row > left_max_row
        or right_max_col < left_min_col
        or right_min_col > left_max_col
    )


def _range_overlap(left: str, right: str) -> bool:
    return _bounds_overlap(range_boundaries(normalize_ref(left)), range_boundaries(normalize_ref(right)))
