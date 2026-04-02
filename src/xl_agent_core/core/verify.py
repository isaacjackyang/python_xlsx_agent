from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl.utils.cell import get_column_letter, range_boundaries

from xl_agent_core.core.contracts import (
    DiffEntry,
    DiffResult,
    ProofCellResult,
    ProofResult,
    ProofTargetResult,
    WarningMessage,
)
from xl_agent_core.core.loader import display_value, load_workbook_bundle, normalize_ref


class VerifyService:
    """Proof and diff workflows for comparing original and modified workbooks."""

    def diff(self, original: str, modified: str) -> DiffResult:
        original_bundle = load_workbook_bundle(original)
        modified_bundle = load_workbook_bundle(modified)

        added_sheets = sorted(set(modified_bundle.formulas.sheetnames) - set(original_bundle.formulas.sheetnames))
        removed_sheets = sorted(set(original_bundle.formulas.sheetnames) - set(modified_bundle.formulas.sheetnames))
        warnings: list[WarningMessage] = []
        entries: list[DiffEntry] = []

        for sheet in sorted(set(original_bundle.formulas.sheetnames) & set(modified_bundle.formulas.sheetnames)):
            original_formula_ws = original_bundle.formulas[sheet]
            original_value_ws = original_bundle.values[sheet]
            modified_formula_ws = modified_bundle.formulas[sheet]
            modified_value_ws = modified_bundle.values[sheet]

            refs = self._occupied_refs(original_formula_ws) | self._occupied_refs(modified_formula_ws)
            for ref in sorted(refs, key=self._sort_key):
                before_formula, before_value = self._cell_state(original_formula_ws, original_value_ws, ref)
                after_formula, after_value = self._cell_state(modified_formula_ws, modified_value_ws, ref)
                classification = self._classify_change(before_formula, after_formula, before_value, after_value)
                if classification is None:
                    continue

                entries.append(
                    DiffEntry(
                        sheet=sheet,
                        ref=ref,
                        classification=classification,
                        before_value=before_value,
                        after_value=after_value,
                        before_formula=before_formula,
                        after_formula=after_formula,
                    ),
                )

        summary = dict(Counter(entry.classification for entry in entries))
        if added_sheets:
            warnings.append(
                WarningMessage(
                    code="added_sheets",
                    message=f"Modified workbook contains added sheets: {added_sheets}",
                ),
            )
        if removed_sheets:
            warnings.append(
                WarningMessage(
                    code="removed_sheets",
                    message=f"Modified workbook is missing sheets from the original: {removed_sheets}",
                ),
            )

        return DiffResult(
            original_path=str(original_bundle.path),
            modified_path=str(modified_bundle.path),
            changed_cells=len(entries),
            summary=summary,
            entries=entries,
            added_sheets=added_sheets,
            removed_sheets=removed_sheets,
            status="ok",
            warnings=warnings,
        )

    def proof(self, baseline: str, current: str, targets: list[str]) -> ProofResult:
        baseline_bundle = load_workbook_bundle(baseline)
        current_bundle = load_workbook_bundle(current)
        proof_targets: list[ProofTargetResult] = []
        warnings: list[WarningMessage] = []
        compared_cell_count = 0
        changed_cell_count = 0

        for target in targets:
            parsed_sheet, parsed_ref = self._parse_target(target)
            cells: list[ProofCellResult] = []
            target_warnings: list[WarningMessage] = []

            if parsed_sheet not in baseline_bundle.formulas.sheetnames:
                target_warnings.append(
                    WarningMessage(
                        code="missing_baseline_sheet",
                        message=f"Baseline workbook does not contain sheet '{parsed_sheet}'.",
                    ),
                )
            if parsed_sheet not in current_bundle.formulas.sheetnames:
                target_warnings.append(
                    WarningMessage(
                        code="missing_current_sheet",
                        message=f"Current workbook does not contain sheet '{parsed_sheet}'.",
                    ),
                )

            if target_warnings:
                proof_targets.append(
                    ProofTargetResult(
                        target=target,
                        cells=[],
                        changed_cell_count=0,
                        warnings=target_warnings,
                    ),
                )
                warnings.extend(target_warnings)
                continue

            baseline_formula_ws = baseline_bundle.formulas[parsed_sheet]
            baseline_value_ws = baseline_bundle.values[parsed_sheet]
            current_formula_ws = current_bundle.formulas[parsed_sheet]
            current_value_ws = current_bundle.values[parsed_sheet]

            for ref in self._expand_target_ref(parsed_ref):
                before_formula, before_value = self._cell_state(baseline_formula_ws, baseline_value_ws, ref)
                after_formula, after_value = self._cell_state(current_formula_ws, current_value_ws, ref)
                classification = self._classify_change(before_formula, after_formula, before_value, after_value)
                changed = classification is not None
                compared_cell_count += 1
                if changed:
                    changed_cell_count += 1

                cells.append(
                    ProofCellResult(
                        sheet=parsed_sheet,
                        ref=ref,
                        classification=classification,
                        before_value=before_value,
                        after_value=after_value,
                        before_formula=before_formula,
                        after_formula=after_formula,
                        changed=changed,
                    ),
                )

            proof_targets.append(
                ProofTargetResult(
                    target=target,
                    cells=cells,
                    changed_cell_count=sum(1 for cell in cells if cell.changed),
                    warnings=target_warnings,
                ),
            )

        return ProofResult(
            baseline_path=str(baseline_bundle.path),
            current_path=str(current_bundle.path),
            targets=proof_targets,
            compared_cell_count=compared_cell_count,
            changed_cell_count=changed_cell_count,
            status="ok",
            warnings=warnings,
        )

    def _occupied_refs(self, worksheet: Any) -> set[str]:
        refs: set[str] = set()
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    refs.add(cell.coordinate)
        return refs

    def _cell_state(self, formula_ws: Any, value_ws: Any, ref: str) -> tuple[str | None, Any]:
        formula_cell = formula_ws[ref]
        value_cell = value_ws[ref]
        formula = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None
        value = display_value(formula_cell, value_cell)
        return formula, value

    def _classify_change(
        self,
        before_formula: str | None,
        after_formula: str | None,
        before_value: Any,
        after_value: Any,
    ) -> str | None:
        if before_formula != after_formula:
            if before_formula and after_formula:
                return "formula_changed"
            if before_formula and not after_formula:
                return "formula_removed"
            if after_formula and not before_formula:
                return "formula_added"

        if before_value != after_value:
            if before_formula and after_formula and before_formula == after_formula:
                return "recalc_impact"
            if before_value is None and after_value is not None:
                return "value_added"
            if before_value is not None and after_value is None:
                return "value_cleared"
            return "value_changed"

        return None

    def _parse_target(self, target: str) -> tuple[str, str]:
        if "!" not in target:
            raise ValueError(f"Target '{target}' must use Sheet!Ref format.")
        sheet, ref = target.split("!", 1)
        return sheet, normalize_ref(ref)

    def _expand_target_ref(self, ref: str) -> list[str]:
        normalized = normalize_ref(ref)
        if ":" not in normalized:
            return [normalized]

        min_col, min_row, max_col, max_row = range_boundaries(normalized)
        refs: list[str] = []
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                refs.append(f"{get_column_letter(column)}{row}")
        return refs

    def _sort_key(self, ref: str) -> tuple[int, int]:
        normalized = normalize_ref(ref)
        for index, char in enumerate(normalized):
            if char.isdigit():
                column_name = normalized[:index]
                row_number = normalized[index:]
                from openpyxl.utils.cell import column_index_from_string

                return int(row_number), column_index_from_string(column_name)
        return 0, 0
