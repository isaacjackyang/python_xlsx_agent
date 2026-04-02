from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from xl_agent_core.core.contracts import (
    MutationEdit,
    MutationPlan,
    MutationResult,
    WarningMessage,
    WorkbookCopyResult,
)
from xl_agent_core.core.loader import coordinate_to_tuple, get_sheet_pair, load_workbook_bundle, normalize_ref
from xl_agent_core.core.recalc import RecalcService


class MutationService:
    """Safe mutation helpers built around copy-on-write workflows."""

    def __init__(self) -> None:
        self._recalc = RecalcService()

    def copy_workbook(self, source_path: str, output_path: str, *, overwrite: bool = False) -> WorkbookCopyResult:
        source = Path(source_path).expanduser().resolve()
        destination = Path(output_path).expanduser().resolve()
        if not source.exists():
            raise FileNotFoundError(f"Workbook was not found: {source}")
        if destination.exists() and not overwrite:
            raise FileExistsError(f"Output workbook already exists: {destination}")

        destination.parent.mkdir(parents=True, exist_ok=True)
        destination.write_bytes(source.read_bytes())

        warnings = self._roundtrip_warnings(source)
        return WorkbookCopyResult(
            workbook_path=str(destination),
            source_path=str(source),
            output_path=str(destination),
            copied=True,
            overwrite=overwrite,
            warnings=warnings,
        )

    def plan_write(
        self,
        path: str,
        sheet: str,
        edits: dict[str, Any],
        *,
        output_path: str | None = None,
        overwrite: bool = False,
        operation: str = "write_cells",
        edit_kind: str = "value",
        dry_run: bool = True,
    ) -> MutationPlan:
        bundle = load_workbook_bundle(path)
        formula_ws, value_ws = get_sheet_pair(bundle, sheet)
        resolved_output = str(Path(output_path).expanduser().resolve()) if output_path else None
        warnings = self._roundtrip_warnings(bundle.path)
        mutation_edits: list[MutationEdit] = []

        for ref, new_value in sorted(edits.items(), key=lambda item: coordinate_to_tuple(item[0])):
            normalized_ref = normalize_ref(ref)
            formula_cell = formula_ws[normalized_ref]
            value_cell = value_ws[normalized_ref]
            before_formula = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None
            before_value = value_cell.value if value_cell.value is not None else formula_cell.value

            if edit_kind == "formula":
                after_formula = self._normalize_formula(new_value)
                after_value = after_formula
                note = "replaces_existing_formula" if before_formula else "sets_formula"
            else:
                after_formula = None
                after_value = new_value
                if before_formula:
                    note = "overwrites_existing_formula"
                else:
                    note = "writes_literal_value"

            mutation_edits.append(
                MutationEdit(
                    sheet=sheet,
                    ref=normalized_ref,
                    kind=edit_kind,
                    before_value=before_value,
                    after_value=after_value,
                    before_formula=before_formula,
                    after_formula=after_formula,
                    note=note,
                ),
            )

            if formula_ws.row_dimensions[formula_cell.row].hidden:
                warnings.append(
                    WarningMessage(
                        code="target_row_hidden",
                        message=f"Target {sheet}!{normalized_ref} is on a hidden row.",
                    ),
                )
            if formula_ws.column_dimensions[formula_cell.column_letter].hidden:
                warnings.append(
                    WarningMessage(
                        code="target_column_hidden",
                        message=f"Target {sheet}!{normalized_ref} is on a hidden column.",
                    ),
                )
            for merged_range in formula_ws.merged_cells.ranges:
                if normalized_ref in merged_range:
                    warnings.append(
                        WarningMessage(
                            code="target_is_merged",
                            message=f"Target {sheet}!{normalized_ref} belongs to merged range {merged_range.coord}.",
                        ),
                    )
                    break

        return MutationPlan(
            operation=operation,
            workbook_path=str(bundle.path),
            sheet=sheet,
            output_path=resolved_output,
            targets=[edit.ref for edit in mutation_edits],
            edits=mutation_edits,
            dry_run=dry_run,
            copy_on_write=resolved_output is not None and resolved_output != str(bundle.path),
            overwrite=overwrite,
            status="planned",
            warnings=self._dedupe_warnings(warnings),
        )

    def apply_write(
        self,
        path: str,
        sheet: str,
        edits: dict[str, Any],
        *,
        output_path: str,
        overwrite: bool = False,
        recalc_backend: str | None = None,
    ) -> MutationResult:
        plan = self.plan_write(
            path,
            sheet,
            edits,
            output_path=output_path,
            overwrite=overwrite,
            operation="write_cells",
            edit_kind="value",
            dry_run=False,
        )
        return self._apply_plan(plan, edit_kind="value", edits=edits, recalc_backend=recalc_backend)

    def replace_formulas(
        self,
        path: str,
        sheet: str,
        edits: dict[str, str],
        *,
        output_path: str,
        overwrite: bool = False,
        recalc_backend: str | None = None,
    ) -> MutationResult:
        plan = self.plan_write(
            path,
            sheet,
            edits,
            output_path=output_path,
            overwrite=overwrite,
            operation="replace_formulas",
            edit_kind="formula",
            dry_run=False,
        )
        return self._apply_plan(plan, edit_kind="formula", edits=edits, recalc_backend=recalc_backend)

    def _apply_plan(
        self,
        plan: MutationPlan,
        *,
        edit_kind: str,
        edits: dict[str, Any],
        recalc_backend: str | None,
    ) -> MutationResult:
        source_path = Path(plan.workbook_path)
        if not plan.output_path:
            raise ValueError("output_path is required for non-destructive mutation workflows.")
        output_path = Path(plan.output_path).expanduser().resolve()
        if output_path.exists() and not plan.overwrite:
            raise FileExistsError(f"Output workbook already exists: {output_path}")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        keep_vba = source_path.suffix.lower() in {".xlsm", ".xltm"}
        workbook = load_workbook(source_path, keep_vba=keep_vba)
        if plan.sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet '{plan.sheet}' was not found.")
        worksheet = workbook[plan.sheet]

        for ref, raw_value in edits.items():
            normalized_ref = normalize_ref(ref)
            if edit_kind == "formula":
                worksheet[normalized_ref] = self._normalize_formula(raw_value)
            else:
                worksheet[normalized_ref] = raw_value

        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcOnSave = True
        workbook.save(output_path)

        recalc_result = None
        warnings = list(plan.warnings)
        if recalc_backend and recalc_backend.lower() != "none":
            recalc_result = self._recalc.recalculate(str(output_path), recalc_backend)
            warnings.extend(recalc_result.warnings)

        return MutationResult(
            workbook_path=str(output_path),
            output_path=str(output_path),
            operation=plan.operation,
            saved=True,
            edit_count=len(plan.edits),
            sheet=plan.sheet,
            edits=plan.edits,
            recalc=recalc_result,
            warnings=self._dedupe_warnings(warnings),
        )

    def _normalize_formula(self, value: Any) -> str:
        text = str(value).strip()
        if not text:
            raise ValueError("Formula replacement cannot be empty.")
        return text if text.startswith("=") else f"={text}"

    def _roundtrip_warnings(self, path: Path) -> list[WarningMessage]:
        warnings = [
            WarningMessage(
                code="openpyxl_roundtrip",
                message=(
                    "Workbook writes are performed through openpyxl. Advanced Excel features may need "
                    "extra validation after save."
                ),
            ),
        ]
        if path.suffix.lower() in {".xlsm", ".xltm"}:
            warnings.append(
                WarningMessage(
                    code="macro_workbook",
                    message="Macro-enabled workbook detected. Validate VBA-related behavior after mutation.",
                ),
            )
        return warnings

    def _dedupe_warnings(self, warnings: list[WarningMessage]) -> list[WarningMessage]:
        seen: set[tuple[str, str]] = set()
        output: list[WarningMessage] = []
        for warning in warnings:
            key = (warning.code, warning.message)
            if key in seen:
                continue
            seen.add(key)
            output.append(warning)
        return output
